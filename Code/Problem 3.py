import math
import argparse
import random
from typing import Tuple, List

Vec3 = Tuple[float, float, float]

# ---------------- Vector utilities ----------------

def dot(a: Vec3, b: Vec3) -> float:
    return a[0]*b[0] + a[1]*b[1] + a[2]*b[2]


def sub(a: Vec3, b: Vec3) -> Vec3:
    return (a[0]-b[0], a[1]-b[1], a[2]-b[2])


def add(a: Vec3, b: Vec3) -> Vec3:
    return (a[0]+b[0], a[1]+b[1], a[2]+b[2])


def mul(a: Vec3, k: float) -> Vec3:
    return (a[0]*k, a[1]*k, a[2]*k)


def norm(a: Vec3) -> float:
    return math.sqrt(dot(a, a))


def clamp(x: float, lo: float, hi: float) -> float:
    return max(lo, min(hi, x))

# ---------------- Geometry helpers ----------------

def dist_point_to_segment(P: Vec3, A: Vec3, B: Vec3) -> float:
    """Distance from point P to segment AB in 3D."""
    AB = sub(B, A)
    AP = sub(P, A)
    ab2 = dot(AB, AB)
    if ab2 == 0.0:
        return norm(sub(P, A))
    t = clamp(dot(AP, AB) / ab2, 0.0, 1.0)
    closest = add(A, mul(AB, t))
    return norm(sub(P, closest))


def cloud_center_after(Pe: Vec3, t: float, t_explosion: float, sink_speed: float = 3.0) -> Vec3:
    """Cloud center position at time t >= t_explosion. Only z changes (downwards)."""
    dt = max(0.0, t - t_explosion)
    return (Pe[0], Pe[1], Pe[2] - sink_speed*dt)


def effective_obscuration_time_for_M1(speed: float,
                                      heading_deg: float,
                                      t_release: float,
                                      t_fuse: float,
                                      g: float = 9.8,
                                      dt: float = 0.001) -> float:
    """
    Compute total effective obscuration time (seconds) for M1 using given parameters.
    Criterion: distance from cloud center to segment [M(t), T] <= 10 m within 20 s after explosion.
    dt: simulation step size in seconds (use a coarser dt for faster optimization, e.g., 0.02; refine later with 0.001).
    """
    # Target (true target) center of bottom face
    T = (0.0, 200.0, 0.0)

    # Compute explosion point and time
    Pe, t_explosion, _ = compute_explosion_point_FY1_new_standard(speed, heading_deg, t_release, t_fuse, g)

    # Simulation window
    t0 = t_explosion
    t1 = t_explosion + 20.0

    # Numerical integration
    R_eff = 10.0

    t = t0
    prev_d = None
    total = 0.0

    while t <= t1 + 1e-12:
        C = cloud_center_after(Pe, t, t_explosion, sink_speed=3.0)
        M = missile_state_M1(t)
        d = dist_point_to_segment(C, M, T)

        if prev_d is not None:
            if (prev_d <= R_eff and d <= R_eff):
                total += dt
            elif (prev_d <= R_eff and d > R_eff):
                denom = (d - prev_d)
                if abs(denom) > 1e-12:
                    alpha = (R_eff - prev_d) / denom
                    if 0.0 < alpha < 1.0:
                        total += alpha * dt
            elif (prev_d > R_eff and d <= R_eff):
                denom = (d - prev_d)
                if abs(denom) > 1e-12:
                    alpha = (R_eff - prev_d) / denom
                    if 0.0 < alpha < 1.0:
                        total += (1.0 - alpha) * dt
                else:
                    total += dt
        prev_d = d
        t += dt

    return total

# ---------------- Optimization: Particle Swarm Optimization (PSO) ----------------

def wrap_angle_deg(a: float) -> float:
    """Wrap angle to [0, 360]."""
    a = a % 360.0
    if a < 0:
        a += 360.0
    return a


def shortest_ang_diff_deg(target: float, current: float) -> float:
    """Return signed shortest angular difference (target - current) in degrees within [-180, 180]."""
    d = (target - current + 180.0) % 360.0 - 180.0
    return d


def pso_optimize(iters: int = 150,
                 swarm_size: int = 25,
                 seed: int = 42,
                 speed_min: float = 70.0,
                 speed_max: float = 140.0,
                 heading_min: float = 0.0,
                 heading_max: float = 360.0,
                 t_release_min: float = 1e-3,
                 t_release_max: float = 3.0,
                 t_fuse_min: float = 1e-3,
                 t_fuse_max: float = 1.5,
                 w: float = 0.7,
                 c1: float = 1.5,
                 c2: float = 1.5,
                 dt_coarse: float = 0.03) -> Tuple[Tuple[float, float, float, float], float]:
    """
    Particle Swarm Optimization to maximize obscuration time.
    Returns (best_params_tuple, best_value), where params = (speed, heading_deg, t_release, t_fuse).
    Runtime is bounded by swarm_size * iters objective evaluations with coarse dt.
    """
    random.seed(seed)

    # Ranges for clamping and velocity limits
    rng_speed = speed_max - speed_min
    rng_heading = heading_max - heading_min  # assumed 360
    rng_tr = t_release_max - t_release_min
    rng_tf = t_fuse_max - t_fuse_min

    # Velocity limits (keep stable)
    vmax = [rng_speed * 0.25, 90.0, rng_tr * 0.25, rng_tf * 0.25]

    def clamp_param(x, lo, hi):
        return max(lo, min(hi, x))

    def obj(xx):
        return effective_obscuration_time_for_M1(xx[0], xx[1], xx[2], xx[3], dt=dt_coarse)

    # Initialize swarm
    X = []  # positions
    V = []  # velocities
    Pbest = []  # personal best positions
    PbestVal = []

    Gbest = None
    GbestVal = -1e18

    for _ in range(swarm_size):
        s = random.uniform(speed_min, speed_max)
        h = random.uniform(heading_min, heading_max)
        tr = random.uniform(t_release_min, t_release_max)
        tf = random.uniform(t_fuse_min, t_fuse_max)
        x = [s, h, tr, tf]
        # Small initial velocities
        v = [random.uniform(-rng_speed*0.05, rng_speed*0.05),
             random.uniform(-30.0, 30.0),
             random.uniform(-rng_tr*0.05, rng_tr*0.05),
             random.uniform(-rng_tf*0.05, rng_tf*0.05)]
        X.append(x)
        V.append(v)
        val = obj(x)
        Pbest.append(x[:])
        PbestVal.append(val)
        if val > GbestVal:
            GbestVal = val
            Gbest = x[:]

    # PSO main loop
    for it in range(iters):
        for i in range(swarm_size):
            x = X[i]
            v = V[i]

            # Update velocity per dimension
            # speed
            r1 = random.random(); r2 = random.random()
            v0 = (w * v[0]
                  + c1 * r1 * (Pbest[i][0] - x[0])
                  + c2 * r2 * (Gbest[0] - x[0]))
            # heading (circular distance)
            r1 = random.random(); r2 = random.random()
            diff_p = shortest_ang_diff_deg(Pbest[i][1], x[1])
            diff_g = shortest_ang_diff_deg(Gbest[1], x[1])
            v1 = (w * v[1]
                  + c1 * r1 * diff_p
                  + c2 * r2 * diff_g)
            # t_release
            r1 = random.random(); r2 = random.random()
            v2 = (w * v[2]
                  + c1 * r1 * (Pbest[i][2] - x[2])
                  + c2 * r2 * (Gbest[2] - x[2]))
            # t_fuse
            r1 = random.random(); r2 = random.random()
            v3 = (w * v[3]
                  + c1 * r1 * (Pbest[i][3] - x[3])
                  + c2 * r2 * (Gbest[3] - x[3]))

            # Velocity clamp
            v0 = max(-vmax[0], min(vmax[0], v0))
            v1 = max(-vmax[1], min(vmax[1], v1))
            v2 = max(-vmax[2], min(vmax[2], v2))
            v3 = max(-vmax[3], min(vmax[3], v3))
            v = [v0, v1, v2, v3]

            # Update position
            x0 = clamp_param(x[0] + v0, speed_min, speed_max)
            x1 = wrap_angle_deg(x[1] + v1)
            x2 = clamp_param(x[2] + v2, t_release_min, t_release_max)
            x3 = clamp_param(x[3] + v3, t_fuse_min, t_fuse_max)
            x = [x0, x1, x2, x3]

            # Save back
            X[i] = x
            V[i] = v

            # Evaluate and update bests
            val = obj(x)
            if val > PbestVal[i]:
                Pbest[i] = x[:]
                PbestVal[i] = val
                if val > GbestVal:
                    GbestVal = val
                    Gbest = x[:]

    return (Gbest[0], Gbest[1], Gbest[2], Gbest[3]), GbestVal


def heading_to_unit_new(theta_deg: float) -> Tuple[float, float]:
    """New standard: 0 deg along +x, CCW positive (xy-plane). Returns (ux, uy)."""
    th = math.radians(theta_deg)
    return (math.cos(th), math.sin(th))


def missile_state_M1(t: float) -> Vec3:
    """M1 position at time t [s]. Missile heads towards (0,0,0) at 300 m/s."""
    M0 = (20000.0, 0.0, 2000.0)
    v_dir = (-20000.0, 0.0, -2000.0)
    L = math.sqrt(v_dir[0]**2 + v_dir[1]**2 + v_dir[2]**2)
    u = (v_dir[0]/L, v_dir[1]/L, v_dir[2]/L)
    v = (u[0]*300.0, u[1]*300.0, u[2]*300.0)
    return (M0[0] + v[0]*t, M0[1] + v[1]*t, M0[2] + v[2]*t)


def compute_explosion_point_FY1_new_standard(speed: float = 132.86,
                                             heading_deg: float = 6.63,
                                             t_release: float = 0.12,
                                             t_fuse: float = 0.68,
                                             g: float = 9.8) -> Tuple[Vec3, float, Tuple[float, float]]:
    """
    Compute FY1 explosion point using the NEW heading standard directly.
    Returns (Pe, t_explosion, (ux, uy)).
    """
    FY1 = (17800.0, 0.0, 1800.0)

    ux, uy = heading_to_unit_new(heading_deg)

    # Drone horizontal velocity components
    vx = speed * ux
    vy = speed * uy

    # Release point (constant altitude prior to release)
    R = (FY1[0] + vx*t_release, FY1[1] + vy*t_release, FY1[2])

    # Free fall after release for t_fuse seconds: horizontal continues, vertical under gravity
    x_e = R[0] + vx*t_fuse
    y_e = R[1] + vy*t_fuse
    z_e = R[2] - 0.5*g*(t_fuse**2)

    Pe = (x_e, y_e, z_e)
    t_explosion = t_release + t_fuse
    return Pe, t_explosion, (ux, uy)


def cloud_center_after(Pe: Vec3, t: float, t_explosion: float, sink_speed: float = 3.0) -> Vec3:
    """Cloud center position at time t >= t_explosion. Only z changes (downwards)."""
    dt = max(0.0, t - t_explosion)
    return (Pe[0], Pe[1], Pe[2] - sink_speed*dt)


def effective_obscuration_time_multi(speed: float,
                                     heading_deg: float,
                                     t_releases: List[float],
                                     t_fuses: List[float],
                                     g: float = 9.8,
                                     dt: float = 0.02) -> float:
    """Union-of-clouds obscuration time for multiple bombs (M1 vs T)."""
    assert len(t_releases) == len(t_fuses)
    # Compute all explosion points and times
    Pe_list: List[Vec3] = []
    Te_list: List[float] = []
    for tr, tf in zip(t_releases, t_fuses):
        Pe, t_explosion, _ = compute_explosion_point_FY1_new_standard(speed, heading_deg, tr, tf, g)
        Pe_list.append(Pe)
        Te_list.append(t_explosion)
    # Time window across all bombs
    t0 = min(Te_list)
    t1 = max(Te_list) + 20.0
    R_eff = 10.0
    T = (0.0, 200.0, 0.0)
    total = 0.0
    prev_min_d = None
    t = t0
    while t <= t1 + 1e-12:
        # Compute current min distance among active clouds
        min_d = float('inf')
        for Pe, te in zip(Pe_list, Te_list):
            if t < te or t > te + 20.0:
                continue
            C = cloud_center_after(Pe, t, te, sink_speed=3.0)
            M = missile_state_M1(t)
            d = dist_point_to_segment(C, M, T)
            if d < min_d:
                min_d = d
        if min_d == float('inf'):
            min_d = 1e9  # no active clouds
        if prev_min_d is not None:
            if (prev_min_d <= R_eff and min_d <= R_eff):
                total += dt
            elif (prev_min_d <= R_eff and min_d > R_eff):
                denom = (min_d - prev_min_d)
                if abs(denom) > 1e-12:
                    alpha = (R_eff - prev_min_d) / denom
                    if 0.0 < alpha < 1.0:
                        total += alpha * dt
            elif (prev_min_d > R_eff and min_d <= R_eff):
                denom = (min_d - prev_min_d)
                if abs(denom) > 1e-12:
                    alpha = (R_eff - prev_min_d) / denom
                    if 0.0 < alpha < 1.0:
                        total += (1.0 - alpha) * dt
                else:
                    total += dt
        prev_min_d = min_d
        t += dt
    return total


def compute_explosions_FY1_multi(speed: float,
                                  heading_deg: float,
                                  t_releases: List[float],
                                  t_fuses: List[float],
                                  g: float = 9.8) -> Tuple[List[Vec3], List[float]]:
    Pe_list: List[Vec3] = []
    Te_list: List[float] = []
    for tr, tf in zip(t_releases, t_fuses):
        Pe, te, _ = compute_explosion_point_FY1_new_standard(speed, heading_deg, tr, tf, g)
        Pe_list.append(Pe)
        Te_list.append(te)
    return Pe_list, Te_list


def pso_optimize_multi(iters: int = 150,
                       swarm_size: int = 30,
                       seed: int = 42,
                       num_bombs: int = 3,
                       speed_min: float = 70.0,
                       speed_max: float = 140.0,
                       heading_min: float = 0.0,
                       heading_max: float = 360.0,
                       t_release_min: float = 1e-3,
                       t_release_max: float = 3.0,
                       t_fuse_min: float = 1e-3,
                       t_fuse_max: float = 1.5,
                       spacing_min: float = 1.0,
                       w: float = 0.7,
                       c1: float = 1.5,
                       c2: float = 1.5,
                       dt_coarse: float = 0.04,
                       require_all_positive: bool = False,
                       min_effective: float = 0.0,
                       penalty_weight: float = 5.0) -> Tuple[Tuple[float, float, List[float], List[float]], float]:
    """PSO for multi-bomb plan (speed, heading, releases, fuses)."""
    assert num_bombs >= 1
    random.seed(seed)

    def project_params(x: List[float]) -> List[float]:
        # x = [speed, heading, r1..rk, f1..fk]
        speed = clamp(x[0], speed_min, speed_max)
        heading = wrap_angle_deg(x[1])
        r = x[2:2+num_bombs]
        f = x[2+num_bombs:2+2*num_bombs]
        # sort and enforce spacing >= spacing_min
        r = sorted(r)
        # Ensure feasible window exists
        min_span = spacing_min*(num_bombs-1)
        if t_release_max - t_release_min < min_span:
            base = t_release_min
        else:
            # Clamp first so that all can fit
            if abs(t_release_min) < 1e-12:
                base = 0.0
            else:
                base = clamp(r[0], t_release_min, t_release_max - min_span)
        r[0] = base
        for i in range(1, num_bombs):
            r[i] = max(r[i], r[i-1] + spacing_min)
        # If last exceeds max, shift left
        overflow = r[-1] - t_release_max
        if overflow > 0:
            for i in range(num_bombs-1, -1, -1):
                r[i] -= overflow
                if i>0 and r[i] < r[i-1] + spacing_min:
                    r[i-1] = r[i] - spacing_min
            # Ensure first >= min
            if r[0] < t_release_min:
                delta = t_release_min - r[0]
                for i in range(num_bombs):
                    r[i] += delta
        # Clamp fuses
        f = [clamp(v, t_fuse_min, t_fuse_max) for v in f]
        return [speed, heading] + r + f

    def obj(xvec: List[float]) -> float:
        xvec = project_params(xvec)
        speed = xvec[0]; heading = xvec[1]
        r = xvec[2:2+num_bombs]
        f = xvec[2+num_bombs:2+2*num_bombs]
        # 带惩罚的目标：联合时长减去对每枚弹的不足惩罚
        union_time = effective_obscuration_time_multi(speed, heading, r, f, dt=dt_coarse)
        if require_all_positive or min_effective > 0.0:
            per = [effective_obscuration_time_for_M1(speed, heading, r[i], f[i], dt=dt_coarse) for i in range(num_bombs)]
            penalty = 0.0
            for t in per:
                if t < min_effective:
                    penalty += (min_effective - t)
            return union_time - penalty_weight * penalty
        else:
            return union_time

    # Ranges for velocity scale
    rng_speed = speed_max - speed_min
    rng_tr = max(1e-6, t_release_max - t_release_min)
    rng_tf = max(1e-6, t_fuse_max - t_fuse_min)

    dims = 2 + 2*num_bombs
    # Initialize swarm
    X: List[List[float]] = []
    V: List[List[float]] = []
    Pbest: List[List[float]] = []
    PbestVal: List[float] = []
    Gbest: List[float] = []
    GbestVal = -1e18

    for _ in range(swarm_size):
        s = random.uniform(speed_min, speed_max)
        h = random.uniform(heading_min, heading_max)
        r = [random.uniform(t_release_min, t_release_max) for _ in range(num_bombs)]
        f = [random.uniform(t_fuse_min, t_fuse_max) for _ in range(num_bombs)]
        x0 = [s, h] + r + f
        x0 = project_params(x0)
        v0 = [
            random.uniform(-rng_speed*0.05, rng_speed*0.05),  # speed
            random.uniform(-30.0, 30.0),                      # heading
        ] + [random.uniform(-rng_tr*0.05, rng_tr*0.05) for _ in range(num_bombs)] \
          + [random.uniform(-rng_tf*0.05, rng_tf*0.05) for _ in range(num_bombs)]
        X.append(x0)
        V.append(v0)
        val = obj(x0)
        Pbest.append(x0[:])
        PbestVal.append(val)
        if val > GbestVal:
            GbestVal = val
            Gbest = x0[:]

    vmax_speed = rng_speed * 0.25
    vmax_heading = 90.0
    vmax_tr = rng_tr * 0.25
    vmax_tf = rng_tf * 0.25

    for it in range(iters):
        for i in range(swarm_size):
            x = X[i][:]
            v = V[i][:]
            # Update each dimension
            # speed
            r1 = random.random(); r2 = random.random()
            v[0] = (w*v[0] + c1*r1*(Pbest[i][0]-x[0]) + c2*r2*(Gbest[0]-x[0]))
            # heading with angular diffs
            r1 = random.random(); r2 = random.random()
            diff_p = shortest_ang_diff_deg(Pbest[i][1], x[1])
            diff_g = shortest_ang_diff_deg(Gbest[1], x[1])
            v[1] = (w*v[1] + c1*r1*diff_p + c2*r2*diff_g)
            # releases
            for d in range(2, 2+num_bombs):
                r1 = random.random(); r2 = random.random()
                v[d] = (w*v[d] + c1*r1*(Pbest[i][d]-x[d]) + c2*r2*(Gbest[d]-x[d]))
            # fuses
            for d in range(2+num_bombs, 2+2*num_bombs):
                r1 = random.random(); r2 = random.random()
                v[d] = (w*v[d] + c1*r1*(Pbest[i][d]-x[d]) + c2*r2*(Gbest[d]-x[d]))
            # Clamp velocities
            v[0] = clamp(v[0], -vmax_speed, vmax_speed)
            v[1] = clamp(v[1], -vmax_heading, vmax_heading)
            for d in range(2, 2+num_bombs):
                v[d] = clamp(v[d], -vmax_tr, vmax_tr)
            for d in range(2+num_bombs, 2+2*num_bombs):
                v[d] = clamp(v[d], -vmax_tf, vmax_tf)
            # Update position and project
            x = [x[d] + v[d] for d in range(dims)]
            x = project_params(x)
            X[i] = x
            V[i] = v
            # Evaluate and update bests
            val = obj(x)
            if val > PbestVal[i]:
                Pbest[i] = x[:]
                PbestVal[i] = val
                if val > GbestVal:
                    GbestVal = val
                    Gbest = x[:]
    # Unpack best
    best_speed = Gbest[0]
    best_heading = Gbest[1]
    best_releases = Gbest[2:2+num_bombs]
    best_fuses = Gbest[2+num_bombs:2+2*num_bombs]
    return (best_speed, best_heading, best_releases, best_fuses), GbestVal


def write_strategy_to_excel(xlsx_path: str,
                            speed: float,
                            heading: float,
                            t_releases: List[float],
                            t_fuses: List[float],
                            Pe_list: List[Vec3],
                            Te_list: List[float],
                            total_time: float,
                            sheet_name: str = None) -> None:
    """Append a new sheet with plan into the given Excel file (template)."""
    try:
        from openpyxl import load_workbook
        from openpyxl.workbook.workbook import Workbook
    except ImportError:
        raise RuntimeError("需要 openpyxl 库来写入 Excel：请先安装 pip install openpyxl")

    wb = load_workbook(xlsx_path)
    if sheet_name is None:
        import datetime
        sheet_name = f"Plan3_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"
    ws = wb.create_sheet(title=sheet_name)

    headers = [
        "无人机", "速度(m/s)", "航向(°)",
        "t_release1(s)", "t_fuse1(s)",
        "t_release2(s)", "t_fuse2(s)",
        "t_release3(s)", "t_fuse3(s)",
        "t_explosion1(s)", "Pe1_x", "Pe1_y", "Pe1_z",
        "t_explosion2(s)", "Pe2_x", "Pe2_y", "Pe2_z",
        "t_explosion3(s)", "Pe3_x", "Pe3_y", "Pe3_z",
        "总遮蔽时长(s)"
    ]
    ws.append(headers)

    row = [
        "FY1", f"{speed:.6f}", f"{heading:.6f}",
        f"{t_releases[0]:.6f}", f"{t_fuses[0]:.6f}",
        f"{t_releases[1]:.6f}", f"{t_fuses[1]:.6f}",
        f"{t_releases[2]:.6f}", f"{t_fuses[2]:.6f}",
        f"{Te_list[0]:.6f}", f"{Pe_list[0][0]:.3f}", f"{Pe_list[0][1]:.3f}", f"{Pe_list[0][2]:.3f}",
        f"{Te_list[1]:.6f}", f"{Pe_list[1][0]:.3f}", f"{Pe_list[1][1]:.3f}", f"{Pe_list[1][2]:.3f}",
        f"{Te_list[2]:.6f}", f"{Pe_list[2][0]:.3f}", f"{Pe_list[2][1]:.3f}", f"{Pe_list[2][2]:.3f}",
        f"{total_time:.6f}"
    ]
    ws.append(row)

    # Add notes
    ws.append([])
    ws.append(["说明：同一无人机连续两次投放间隔 ≥ 1 s；云团下沉 3 m/s，有效期 20 s；判据：距[M1(t),T]线段 ≤ 10m 计为有效遮蔽。"])

    wb.save(xlsx_path)


def main():
    parser = argparse.ArgumentParser(description="Compute or optimize obscuration time for M1 (new heading standard)")
    subparsers = parser.add_subparsers(dest="cmd", required=True)

    # Subcommand: direct evaluation
    p_eval = subparsers.add_parser("eval", help="Direct evaluation with explicit parameters")
    p_eval.add_argument("--speed", type=float, required=True, help="UAV speed in m/s, must be within [70, 140]")
    p_eval.add_argument("--heading", type=float, required=True, help="UAV heading in degrees [0, 360], new standard: 0° along +x, CCW positive")
    p_eval.add_argument("--t_release", type=float, required=True, help="Release time after task received (s), >0")
    p_eval.add_argument("--t_fuse", type=float, required=True, help="Fuse time after release until explosion (s), >0")
    p_eval.add_argument("--dt", type=float, default=0.001, help="Simulation step size in seconds (default 0.001)")

    # Subcommand: optimization via PSO
    p_opt = subparsers.add_parser("opt", help="Optimize parameters using Particle Swarm Optimization (PSO)")
    p_opt.add_argument("--iters", type=int, default=150, help="Number of PSO iterations")
    p_opt.add_argument("--swarm", type=int, default=25, help="Swarm size (number of particles)")
    p_opt.add_argument("--seed", type=int, default=42, help="Random seed")
    p_opt.add_argument("--speed_min", type=float, default=70.0)
    p_opt.add_argument("--speed_max", type=float, default=140.0)
    p_opt.add_argument("--heading_min", type=float, default=0.0)
    p_opt.add_argument("--heading_max", type=float, default=360.0)
    p_opt.add_argument("--t_release_max", type=float, default=3.0)
    p_opt.add_argument("--t_fuse_max", type=float, default=1.5)
    p_opt.add_argument("--w", type=float, default=0.7, help="PSO inertia weight")
    p_opt.add_argument("--c1", type=float, default=1.5, help="PSO cognitive coefficient")
    p_opt.add_argument("--c2", type=float, default=1.5, help="PSO social coefficient")
    p_opt.add_argument("--dt_coarse", type=float, default=0.03, help="Coarse dt for search (e.g., 0.02~0.05)")
    p_opt.add_argument("--refine_dt", type=float, default=0.001, help="Refinement dt for final evaluation (e.g., 0.001~0.002)")
    p_opt.add_argument("--num_bombs", type=int, default=1, help="Number of bombs (1 or 3)")
    p_opt.add_argument("--save_to", type=str, default="", help="Save plan to an Excel file (existing template), e.g., d:/A/result1.xlsx")
    p_opt.add_argument("--spacing_min", type=float, default=2.5, help="Minimum spacing between consecutive releases (s)")
    p_opt.add_argument("--require_all_positive", action="store_true", help="Enforce each bomb to have at least min_effective seconds of effective time via penalty")
    p_opt.add_argument("--min_effective", type=float, default=0.5, help="Per-bomb minimal effective time target (s) when --require_all_positive is set")
    p_opt.add_argument("--penalty_weight", type=float, default=5.0, help="Penalty weight for shortfalls in per-bomb effective time")

    args = parser.parse_args()

    if args.cmd == "eval":
        # Validate ranges
        if not (70.0 <= args.speed <= 140.0):
            raise ValueError(f"speed must be within [70, 140], got {args.speed}")
        if not (0.0 <= args.heading <= 360.0):
            raise ValueError(f"heading must be within [0, 360], got {args.heading}")
        if not (args.t_release > 0.0):
            raise ValueError(f"t_release must be > 0, got {args.t_release}")
        if not (args.t_fuse > 0.0):
            raise ValueError(f"t_fuse must be > 0, got {args.t_fuse}")

        Pe, t_explosion, _ = compute_explosion_point_FY1_new_standard(speed=args.speed,
                                                                      heading_deg=args.heading,
                                                                      t_release=args.t_release,
                                                                      t_fuse=args.t_fuse)
        total_time = effective_obscuration_time_for_M1(speed=args.speed,
                                                       heading_deg=args.heading,
                                                       t_release=args.t_release,
                                                       t_fuse=args.t_fuse,
                                                       dt=args.dt)

        print("采用新方位角定义：0°沿+x，逆时针为正（0~360°）")
        print(f"FY1 航向 = {args.heading:.2f}°（新标准），速度 = {args.speed:.2f} m/s")
        print(f"释放时刻 t_release = {args.t_release:.3f} s，起爆时刻 t_explosion = {t_explosion:.3f} s")
        print(f"起爆点 Pe = ({Pe[0]:.3f}, {Pe[1]:.3f}, {Pe[2]:.3f})")
        print("云团中心以 3 m/s 匀速下沉，有效期 20 s，判据：距离[M1(t), T]线段 ≤ 10 m")
        print(f"对 M1 的有效遮蔽总时长 = {total_time:.6f} s")

    elif args.cmd == "opt":
        if args.num_bombs == 1:
            params, best_val_coarse = pso_optimize(iters=args.iters,
                                                   swarm_size=args.swarm,
                                                   seed=args.seed,
                                                   speed_min=args.speed_min,
                                                   speed_max=args.speed_max,
                                                   heading_min=args.heading_min,
                                                   heading_max=args.heading_max,
                                                   t_release_min=1e-3,
                                                   t_release_max=args.t_release_max,
                                                   t_fuse_min=1e-3,
                                                   t_fuse_max=args.t_fuse_max,
                                                   w=args.w, c1=args.c1, c2=args.c2,
                                                   dt_coarse=args.dt_coarse)
            s, h, tr, tf = params
            # Refine with smaller dt for accurate final time
            final_time = effective_obscuration_time_for_M1(s, h, tr, tf, dt=args.refine_dt)
            Pe, t_explosion, _ = compute_explosion_point_FY1_new_standard(s, h, tr, tf)

            print("优化完成（粒子群算法 PSO，单弹）")
            print(f"参数范围：speed∈[{args.speed_min},{args.speed_max}], heading∈[{args.heading_min},{args.heading_max}], t_release∈(0,{args.t_release_max}], t_fuse∈(0,{args.t_fuse_max}]")
            print(f"粗精度 dt = {args.dt_coarse}，PSO：iters = {args.iters}，swarm = {args.swarm}，seed = {args.seed}")
            print("最优参数（粗评价下）：")
            print(f"  speed = {s:.6f} m/s, heading = {h:.6f} °, t_release = {tr:.6f} s, t_fuse = {tf:.6f} s")
            print(f"粗评价下的遮蔽时长 ~= {best_val_coarse:.6f} s")
            print("细评估（refine）结果：")
            print(f"  起爆时刻 = {t_explosion:.6f} s，起爆点 Pe = ({Pe[0]:.3f}, {Pe[1]:.3f}, {Pe[2]:.3f})")
            print(f"  最终有效遮蔽总时长 = {final_time:.6f} s（dt={args.refine_dt}）")
        elif args.num_bombs == 3:
            (s, h, r_list, f_list), best_val_coarse = pso_optimize_multi(iters=args.iters,
                                                                         swarm_size=max(24, args.swarm),
                                                                         seed=args.seed,
                                                                         num_bombs=3,
                                                                         speed_min=args.speed_min,
                                                                         speed_max=args.speed_max,
                                                                         heading_min=args.heading_min,
                                                                         heading_max=args.heading_max,
                                                                         t_release_min=0.0,
                                                                         t_release_max=args.t_release_max,
                                                                         t_fuse_min=1e-3,
                                                                         t_fuse_max=args.t_fuse_max,
                                                                         spacing_min=args.spacing_min,
                                                                         w=args.w, c1=args.c1, c2=args.c2,
                                                                         dt_coarse=args.dt_coarse,
                                                                         require_all_positive=args.require_all_positive,
                                                                         min_effective=args.min_effective,
                                                                         penalty_weight=args.penalty_weight)
            # Final precise evaluation of union time (for reference)
            union_time = effective_obscuration_time_multi(s, h, r_list, f_list, dt=args.refine_dt)

            # Compute per-bomb details (inline; avoid forward reference)
            R_list = []
            Pe_list = []
            Te_list = []
            per_times = []
            ux, uy = heading_to_unit_new(h)
            vx = s * ux
            vy = s * uy
            FY1 = (17800.0, 0.0, 1800.0)
            for tr, tf in zip(r_list, f_list):
                R = (FY1[0] + vx*tr, FY1[1] + vy*tr, FY1[2])
                x_e = R[0] + vx*tf
                y_e = R[1] + vy*tf
                z_e = R[2] - 0.5*9.8*(tf**2)
                Pe = (x_e, y_e, z_e)
                te = tr + tf
                R_list.append(R)
                Pe_list.append(Pe)
                Te_list.append(te)
                per_t = effective_obscuration_time_for_M1(s, h, tr, tf, dt=args.refine_dt)
                per_times.append(per_t)

            print("优化完成（粒子群算法 PSO，FY1 三弹；同一无人机相邻两弹投放间隔 ≥ 1 s）")
            print(f"FY1 无人机运动方向（°）：{h:.2f}，速度（m/s）：{s:.2f}")
            print(f"PSO 设置：iters = {args.iters}，swarm = {max(24, args.swarm)}，seed = {args.seed}；dt_coarse = {args.dt_coarse}，refine_dt = {args.refine_dt}")
            print("—— 三发分别的投放/起爆与有效干扰时长 ——")
            for i in range(3):
                print(f"第{i+1}枚：")
                print(f"  无人机运动方向（°）：{h:.2f}")
                print(f"  无人机运动速度（m/s）：{s:.2f}")
                print(f"  烟幕干扰弹投放点的x坐标（m）：{R_list[i][0]:.3f}")
                print(f"  烟幕干扰弹投放点的y坐标（m）：{R_list[i][1]:.3f}")
                print(f"  烟幕干扰弹投放点的z坐标（m）：{R_list[i][2]:.3f}")
                print(f"  烟幕干扰弹起爆点的x坐标（m）：{Pe_list[i][0]:.3f}")
                print(f"  烟幕干扰弹起爆点的y坐标（m）：{Pe_list[i][1]:.3f}")
                print(f"  烟幕干扰弹起爆点的z坐标（m）：{Pe_list[i][2]:.3f}")
                print(f"  有效干扰时长（s）：{per_times[i]:.6f}")
            print(f"联合（并集）有效干扰总时长 = {union_time:.6f} s")
        else:
            raise ValueError("--num_bombs 目前仅支持 1 或 3")


if __name__ == "__main__":
    main()


def compute_release_and_explosion_FY1_new_standard(speed: float,
                                                   heading_deg: float,
                                                   t_release: float,
                                                   t_fuse: float,
                                                   g: float = 9.8) -> Tuple[Vec3, Vec3, float, Tuple[float, float]]:
    """Return (R, Pe, t_explosion, (ux,uy)) for given parameters under new heading standard."""
    FY1 = (17800.0, 0.0, 1800.0)
    ux, uy = heading_to_unit_new(heading_deg)
    vx = speed * ux
    vy = speed * uy
    R = (FY1[0] + vx*t_release, FY1[1] + vy*t_release, FY1[2])
    x_e = R[0] + vx*t_fuse
    y_e = R[1] + vy*t_fuse
    z_e = R[2] - 0.5*g*(t_fuse**2)
    Pe = (x_e, y_e, z_e)
    te = t_release + t_fuse
    return R, Pe, te, (ux, uy)


def pso_optimize_multi(iters: int = 150,
                       swarm_size: int = 30,
                       seed: int = 42,
                       num_bombs: int = 3,
                       speed_min: float = 70.0,
                       speed_max: float = 140.0,
                       heading_min: float = 0.0,
                       heading_max: float = 360.0,
                       t_release_min: float = 1e-3,
                       t_release_max: float = 3.0,
                       t_fuse_min: float = 1e-3,
                       t_fuse_max: float = 1.5,
                       spacing_min: float = 1.0,
                       w: float = 0.7,
                       c1: float = 1.5,
                       c2: float = 1.5,
                       dt_coarse: float = 0.04,
                       require_all_positive: bool = False,
                       min_effective: float = 0.0,
                       penalty_weight: float = 5.0) -> Tuple[Tuple[float, float, List[float], List[float]], float]:
    """PSO for multi-bomb plan (speed, heading, releases, fuses)."""
    assert num_bombs >= 1
    random.seed(seed)

    def project_params(x: List[float]) -> List[float]:
        # x = [speed, heading, r1..rk, f1..fk]
        speed = clamp(x[0], speed_min, speed_max)
        heading = wrap_angle_deg(x[1])
        r = x[2:2+num_bombs]
        f = x[2+num_bombs:2+2*num_bombs]
        # sort and enforce spacing >= spacing_min
        r = sorted(r)
        # Ensure feasible window exists
        min_span = spacing_min*(num_bombs-1)
        if t_release_max - t_release_min < min_span:
            base = t_release_min
        else:
            # Clamp first so that all can fit
            if abs(t_release_min) < 1e-12:
                base = 0.0
            else:
                base = clamp(r[0], t_release_min, t_release_max - min_span)
        r[0] = base
        for i in range(1, num_bombs):
            r[i] = max(r[i], r[i-1] + spacing_min)
        # If last exceeds max, shift left
        overflow = r[-1] - t_release_max
        if overflow > 0:
            for i in range(num_bombs-1, -1, -1):
                r[i] -= overflow
                if i>0 and r[i] < r[i-1] + spacing_min:
                    r[i-1] = r[i] - spacing_min
            # Ensure first >= min
            if r[0] < t_release_min:
                delta = t_release_min - r[0]
                for i in range(num_bombs):
                    r[i] += delta
        # Clamp fuses
        f = [clamp(v, t_fuse_min, t_fuse_max) for v in f]
        return [speed, heading] + r + f

    def obj(xvec: List[float]) -> float:
        xvec = project_params(xvec)
        speed = xvec[0]; heading = xvec[1]
        r = xvec[2:2+num_bombs]
        f = xvec[2+num_bombs:2+2*num_bombs]
        # 带惩罚的目标：联合时长减去对每枚弹的不足惩罚
        union_time = effective_obscuration_time_multi(speed, heading, r, f, dt=dt_coarse)
        if require_all_positive or min_effective > 0.0:
            per = [effective_obscuration_time_for_M1(speed, heading, r[i], f[i], dt=dt_coarse) for i in range(num_bombs)]
            penalty = 0.0
            for t in per:
                if t < min_effective:
                    penalty += (min_effective - t)
            return union_time - penalty_weight * penalty
        else:
            return union_time

    # Ranges for velocity scale
    rng_speed = speed_max - speed_min
    rng_tr = max(1e-6, t_release_max - t_release_min)
    rng_tf = max(1e-6, t_fuse_max - t_fuse_min)

    dims = 2 + 2*num_bombs
    # Initialize swarm
    X: List[List[float]] = []
    V: List[List[float]] = []
    Pbest: List[List[float]] = []
    PbestVal: List[float] = []
    Gbest: List[float] = []
    GbestVal = -1e18

    for _ in range(swarm_size):
        s = random.uniform(speed_min, speed_max)
        h = random.uniform(heading_min, heading_max)
        r = [random.uniform(t_release_min, t_release_max) for _ in range(num_bombs)]
        f = [random.uniform(t_fuse_min, t_fuse_max) for _ in range(num_bombs)]
        x0 = [s, h] + r + f
        x0 = project_params(x0)
        v0 = [
            random.uniform(-rng_speed*0.05, rng_speed*0.05),  # speed
            random.uniform(-30.0, 30.0),                      # heading
        ] + [random.uniform(-rng_tr*0.05, rng_tr*0.05) for _ in range(num_bombs)] \
          + [random.uniform(-rng_tf*0.05, rng_tf*0.05) for _ in range(num_bombs)]
        X.append(x0)
        V.append(v0)
        val = obj(x0)
        Pbest.append(x0[:])
        PbestVal.append(val)
        if val > GbestVal:
            GbestVal = val
            Gbest = x0[:]

    vmax_speed = rng_speed * 0.25
    vmax_heading = 90.0
    vmax_tr = rng_tr * 0.25
    vmax_tf = rng_tf * 0.25

    for it in range(iters):
        for i in range(swarm_size):
            x = X[i][:]
            v = V[i][:]
            # Update each dimension
            # speed
            r1 = random.random(); r2 = random.random()
            v[0] = (w*v[0] + c1*r1*(Pbest[i][0]-x[0]) + c2*r2*(Gbest[0]-x[0]))
            # heading with angular diffs
            r1 = random.random(); r2 = random.random()
            diff_p = shortest_ang_diff_deg(Pbest[i][1], x[1])
            diff_g = shortest_ang_diff_deg(Gbest[1], x[1])
            v[1] = (w*v[1] + c1*r1*diff_p + c2*r2*diff_g)
            # releases
            for d in range(2, 2+num_bombs):
                r1 = random.random(); r2 = random.random()
                v[d] = (w*v[d] + c1*r1*(Pbest[i][d]-x[d]) + c2*r2*(Gbest[d]-x[d]))
            # fuses
            for d in range(2+num_bombs, 2+2*num_bombs):
                r1 = random.random(); r2 = random.random()
                v[d] = (w*v[d] + c1*r1*(Pbest[i][d]-x[d]) + c2*r2*(Gbest[d]-x[d]))
            # Clamp velocities
            v[0] = clamp(v[0], -vmax_speed, vmax_speed)
            v[1] = clamp(v[1], -vmax_heading, vmax_heading)
            for d in range(2, 2+num_bombs):
                v[d] = clamp(v[d], -vmax_tr, vmax_tr)
            for d in range(2+num_bombs, 2+2*num_bombs):
                v[d] = clamp(v[d], -vmax_tf, vmax_tf)
            # Update position and project
            x = [x[d] + v[d] for d in range(dims)]
            x = project_params(x)
            X[i] = x
            V[i] = v
            # Evaluate and update bests
            val = obj(x)
            if val > PbestVal[i]:
                Pbest[i] = x[:]
                PbestVal[i] = val
                if val > GbestVal:
                    GbestVal = val
                    Gbest = x[:]
    # Unpack best
    best_speed = Gbest[0]
    best_heading = Gbest[1]
    best_releases = Gbest[2:2+num_bombs]
    best_fuses = Gbest[2+num_bombs:2+2*num_bombs]
    return (best_speed, best_heading, best_releases, best_fuses), GbestVal


def write_strategy_to_excel(xlsx_path: str,
                            speed: float,
                            heading: float,
                            t_releases: List[float],
                            t_fuses: List[float],
                            Pe_list: List[Vec3],
                            Te_list: List[float],
                            total_time: float,
                            sheet_name: str = None) -> None:
    """Append a new sheet with plan into the given Excel file (template)."""
    try:
        from openpyxl import load_workbook
        from openpyxl.workbook.workbook import Workbook
    except ImportError:
        raise RuntimeError("需要 openpyxl 库来写入 Excel：请先安装 pip install openpyxl")

    wb = load_workbook(xlsx_path)
    if sheet_name is None:
        import datetime
        sheet_name = f"Plan3_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"
    ws = wb.create_sheet(title=sheet_name)

    headers = [
        "无人机", "速度(m/s)", "航向(°)",
        "t_release1(s)", "t_fuse1(s)",
        "t_release2(s)", "t_fuse2(s)",
        "t_release3(s)", "t_fuse3(s)",
        "t_explosion1(s)", "Pe1_x", "Pe1_y", "Pe1_z",
        "t_explosion2(s)", "Pe2_x", "Pe2_y", "Pe2_z",
        "t_explosion3(s)", "Pe3_x", "Pe3_y", "Pe3_z",
        "总遮蔽时长(s)"
    ]
    ws.append(headers)

    row = [
        "FY1", f"{speed:.6f}", f"{heading:.6f}",
        f"{t_releases[0]:.6f}", f"{t_fuses[0]:.6f}",
        f"{t_releases[1]:.6f}", f"{t_fuses[1]:.6f}",
        f"{t_releases[2]:.6f}", f"{t_fuses[2]:.6f}",
        f"{Te_list[0]:.6f}", f"{Pe_list[0][0]:.3f}", f"{Pe_list[0][1]:.3f}", f"{Pe_list[0][2]:.3f}",
        f"{Te_list[1]:.6f}", f"{Pe_list[1][0]:.3f}", f"{Pe_list[1][1]:.3f}", f"{Pe_list[1][2]:.3f}",
        f"{Te_list[2]:.6f}", f"{Pe_list[2][0]:.3f}", f"{Pe_list[2][1]:.3f}", f"{Pe_list[2][2]:.3f}",
        f"{total_time:.6f}"
    ]
    ws.append(row)

    # Add notes
    ws.append([])
    ws.append(["说明：同一无人机连续两次投放间隔 ≥ 1 s；云团下沉 3 m/s，有效期 20 s；判据：距[M1(t),T]线段 ≤ 10m 计为有效遮蔽。"])

    wb.save(xlsx_path)


def main():
    parser = argparse.ArgumentParser(description="Compute or optimize obscuration time for M1 (new heading standard)")
    subparsers = parser.add_subparsers(dest="cmd", required=True)

    # Subcommand: direct evaluation
    p_eval = subparsers.add_parser("eval", help="Direct evaluation with explicit parameters")
    p_eval.add_argument("--speed", type=float, required=True, help="UAV speed in m/s, must be within [70, 140]")
    p_eval.add_argument("--heading", type=float, required=True, help="UAV heading in degrees [0, 360], new standard: 0° along +x, CCW positive")
    p_eval.add_argument("--t_release", type=float, required=True, help="Release time after task received (s), >0")
    p_eval.add_argument("--t_fuse", type=float, required=True, help="Fuse time after release until explosion (s), >0")
    p_eval.add_argument("--dt", type=float, default=0.001, help="Simulation step size in seconds (default 0.001)")

    # Subcommand: optimization via PSO
    p_opt = subparsers.add_parser("opt", help="Optimize parameters using Particle Swarm Optimization (PSO)")
    p_opt.add_argument("--iters", type=int, default=150, help="Number of PSO iterations")
    p_opt.add_argument("--swarm", type=int, default=25, help="Swarm size (number of particles)")
    p_opt.add_argument("--seed", type=int, default=42, help="Random seed")
    p_opt.add_argument("--speed_min", type=float, default=70.0)
    p_opt.add_argument("--speed_max", type=float, default=140.0)
    p_opt.add_argument("--heading_min", type=float, default=0.0)
    p_opt.add_argument("--heading_max", type=float, default=360.0)
    p_opt.add_argument("--t_release_max", type=float, default=3.0)
    p_opt.add_argument("--t_fuse_max", type=float, default=1.5)
    p_opt.add_argument("--w", type=float, default=0.7, help="PSO inertia weight")
    p_opt.add_argument("--c1", type=float, default=1.5, help="PSO cognitive coefficient")
    p_opt.add_argument("--c2", type=float, default=1.5, help="PSO social coefficient")
    p_opt.add_argument("--dt_coarse", type=float, default=0.03, help="Coarse dt for search (e.g., 0.02~0.05)")
    p_opt.add_argument("--refine_dt", type=float, default=0.001, help="Refinement dt for final evaluation (e.g., 0.001~0.002)")
    p_opt.add_argument("--num_bombs", type=int, default=1, help="Number of bombs (1 or 3)")
    p_opt.add_argument("--save_to", type=str, default="", help="Save plan to an Excel file (existing template), e.g., d:/A/result1.xlsx")

    args = parser.parse_args()

    if args.cmd == "eval":
        # Validate ranges
        if not (70.0 <= args.speed <= 140.0):
            raise ValueError(f"speed must be within [70, 140], got {args.speed}")
        if not (0.0 <= args.heading <= 360.0):
            raise ValueError(f"heading must be within [0, 360], got {args.heading}")
        if not (args.t_release > 0.0):
            raise ValueError(f"t_release must be > 0, got {args.t_release}")
        if not (args.t_fuse > 0.0):
            raise ValueError(f"t_fuse must be > 0, got {args.t_fuse}")

        Pe, t_explosion, _ = compute_explosion_point_FY1_new_standard(speed=args.speed,
                                                                      heading_deg=args.heading,
                                                                      t_release=args.t_release,
                                                                      t_fuse=args.t_fuse)
        total_time = effective_obscuration_time_for_M1(speed=args.speed,
                                                       heading_deg=args.heading,
                                                       t_release=args.t_release,
                                                       t_fuse=args.t_fuse,
                                                       dt=args.dt)

        print("采用新方位角定义：0°沿+x，逆时针为正（0~360°）")
        print(f"FY1 航向 = {args.heading:.2f}°（新标准），速度 = {args.speed:.2f} m/s")
        print(f"释放时刻 t_release = {args.t_release:.3f} s，起爆时刻 t_explosion = {t_explosion:.3f} s")
        print(f"起爆点 Pe = ({Pe[0]:.3f}, {Pe[1]:.3f}, {Pe[2]:.3f})")
        print("云团中心以 3 m/s 匀速下沉，有效期 20 s，判据：距离[M1(t), T]线段 ≤ 10 m")
        print(f"对 M1 的有效遮蔽总时长 = {total_time:.6f} s")

    elif args.cmd == "opt":
        params, best_val_coarse = pso_optimize(iters=args.iters,
                                               swarm_size=args.swarm,
                                               seed=args.seed,
                                               speed_min=args.speed_min,
                                               speed_max=args.speed_max,
                                               heading_min=args.heading_min,
                                               heading_max=args.heading_max,
                                               t_release_min=1e-3,
                                               t_release_max=args.t_release_max,
                                               t_fuse_min=1e-3,
                                               t_fuse_max=args.t_fuse_max,
                                               w=args.w, c1=args.c1, c2=args.c2,
                                               dt_coarse=args.dt_coarse)
        s, h, tr, tf = params
        # Refine with smaller dt for accurate final time
        final_time = effective_obscuration_time_for_M1(s, h, tr, tf, dt=args.refine_dt)
        Pe, t_explosion, _ = compute_explosion_point_FY1_new_standard(s, h, tr, tf)

        print("优化完成（粒子群算法 PSO）")
        print(f"参数范围：speed∈[{args.speed_min},{args.speed_max}], heading∈[{args.heading_min},{args.heading_max}], t_release∈(0,{args.t_release_max}], t_fuse∈(0,{args.t_fuse_max}]")
        print(f"粗精度 dt = {args.dt_coarse}，PSO：iters = {args.iters}，swarm = {args.swarm}，seed = {args.seed}")
        print("最优参数（粗评价下）：")
        print(f"  speed = {s:.6f} m/s, heading = {h:.6f} °, t_release = {tr:.6f} s, t_fuse = {tf:.6f} s")
        print(f"粗评价下的遮蔽时长 ~= {best_val_coarse:.6f} s")
        print("细评估（refine）结果：")
        print(f"  起爆时刻 = {t_explosion:.6f} s，起爆点 Pe = ({Pe[0]:.3f}, {Pe[1]:.3f}, {Pe[2]:.3f})")
        print(f"  最终有效遮蔽总时长 = {final_time:.6f} s（dt={args.refine_dt}）")


if __name__ == "__main__":
    main()


def compute_release_and_explosion_FY1_new_standard(speed: float,
                                                   heading_deg: float,
                                                   t_release: float,
                                                   t_fuse: float,
                                                   g: float = 9.8) -> Tuple[Vec3, Vec3, float, Tuple[float, float]]:
    """Return (R, Pe, t_explosion, (ux,uy)) for given parameters under new heading standard."""
    FY1 = (17800.0, 0.0, 1800.0)
    ux, uy = heading_to_unit_new(heading_deg)
    vx = speed * ux
    vy = speed * uy
    R = (FY1[0] + vx*t_release, FY1[1] + vy*t_release, FY1[2])
    x_e = R[0] + vx*t_fuse
    y_e = R[1] + vy*t_fuse
    z_e = R[2] - 0.5*g*(t_fuse**2)
    Pe = (x_e, y_e, z_e)
    te = t_release + t_fuse
    return R, Pe, te, (ux, uy)


def pso_optimize_multi(iters: int = 150,
                       swarm_size: int = 30,
                       seed: int = 42,
                       num_bombs: int = 3,
                       speed_min: float = 70.0,
                       speed_max: float = 140.0,
                       heading_min: float = 0.0,
                       heading_max: float = 360.0,
                       t_release_min: float = 1e-3,
                       t_release_max: float = 3.0,
                       t_fuse_min: float = 1e-3,
                       t_fuse_max: float = 1.5,
                       spacing_min: float = 1.0,
                       w: float = 0.7,
                       c1: float = 1.5,
                       c2: float = 1.5,
                       dt_coarse: float = 0.04,
                       require_all_positive: bool = False,
                       min_effective: float = 0.0,
                       penalty_weight: float = 5.0) -> Tuple[Tuple[float, float, List[float], List[float]], float]:
    """PSO for multi-bomb plan (speed, heading, releases, fuses)."""
    assert num_bombs >= 1
    random.seed(seed)

    def project_params(x: List[float]) -> List[float]:
        # x = [speed, heading, r1..rk, f1..fk]
        speed = clamp(x[0], speed_min, speed_max)
        heading = wrap_angle_deg(x[1])
        r = x[2:2+num_bombs]
        f = x[2+num_bombs:2+2*num_bombs]
        # sort and enforce spacing >= spacing_min
        r = sorted(r)
        # Ensure feasible window exists
        min_span = spacing_min*(num_bombs-1)
        if t_release_max - t_release_min < min_span:
            base = t_release_min
        else:
            # Clamp first so that all can fit
            if abs(t_release_min) < 1e-12:
                base = 0.0
            else:
                base = clamp(r[0], t_release_min, t_release_max - min_span)
        r[0] = base
        for i in range(1, num_bombs):
            r[i] = max(r[i], r[i-1] + spacing_min)
        # If last exceeds max, shift left
        overflow = r[-1] - t_release_max
        if overflow > 0:
            for i in range(num_bombs-1, -1, -1):
                r[i] -= overflow
                if i>0 and r[i] < r[i-1] + spacing_min:
                    r[i-1] = r[i] - spacing_min
            # Ensure first >= min
            if r[0] < t_release_min:
                delta = t_release_min - r[0]
                for i in range(num_bombs):
                    r[i] += delta
        # Clamp fuses
        f = [clamp(v, t_fuse_min, t_fuse_max) for v in f]
        return [speed, heading] + r + f

    def obj(xvec: List[float]) -> float:
        xvec = project_params(xvec)
        speed = xvec[0]; heading = xvec[1]
        r = xvec[2:2+num_bombs]
        f = xvec[2+num_bombs:2+2*num_bombs]
        # 带惩罚的目标：联合时长减去对每枚弹的不足惩罚
        union_time = effective_obscuration_time_multi(speed, heading, r, f, dt=dt_coarse)
        if require_all_positive or min_effective > 0.0:
            per = [effective_obscuration_time_for_M1(speed, heading, r[i], f[i], dt=dt_coarse) for i in range(num_bombs)]
            penalty = 0.0
            for t in per:
                if t < min_effective:
                    penalty += (min_effective - t)
            return union_time - penalty_weight * penalty
        else:
            return union_time

    # Ranges for velocity scale
    rng_speed = speed_max - speed_min
    rng_tr = max(1e-6, t_release_max - t_release_min)
    rng_tf = max(1e-6, t_fuse_max - t_fuse_min)

    dims = 2 + 2*num_bombs
    # Initialize swarm
    X: List[List[float]] = []
    V: List[List[float]] = []
    Pbest: List[List[float]] = []
    PbestVal: List[float] = []
    Gbest: List[float] = []
    GbestVal = -1e18

    for _ in range(swarm_size):
        s = random.uniform(speed_min, speed_max)
        h = random.uniform(heading_min, heading_max)
        r = [random.uniform(t_release_min, t_release_max) for _ in range(num_bombs)]
        f = [random.uniform(t_fuse_min, t_fuse_max) for _ in range(num_bombs)]
        x0 = [s, h] + r + f
        x0 = project_params(x0)
        v0 = [
            random.uniform(-rng_speed*0.05, rng_speed*0.05),  # speed
            random.uniform(-30.0, 30.0),                      # heading
        ] + [random.uniform(-rng_tr*0.05, rng_tr*0.05) for _ in range(num_bombs)] \
          + [random.uniform(-rng_tf*0.05, rng_tf*0.05) for _ in range(num_bombs)]
        X.append(x0)
        V.append(v0)
        val = obj(x0)
        Pbest.append(x0[:])
        PbestVal.append(val)
        if val > GbestVal:
            GbestVal = val
            Gbest = x0[:]

    vmax_speed = rng_speed * 0.25
    vmax_heading = 90.0
    vmax_tr = rng_tr * 0.25
    vmax_tf = rng_tf * 0.25

    for it in range(iters):
        for i in range(swarm_size):
            x = X[i][:]
            v = V[i][:]
            # Update each dimension
            # speed
            r1 = random.random(); r2 = random.random()
            v[0] = (w*v[0] + c1*r1*(Pbest[i][0]-x[0]) + c2*r2*(Gbest[0]-x[0]))
            # heading with angular diffs
            r1 = random.random(); r2 = random.random()
            diff_p = shortest_ang_diff_deg(Pbest[i][1], x[1])
            diff_g = shortest_ang_diff_deg(Gbest[1], x[1])
            v[1] = (w*v[1] + c1*r1*diff_p + c2*r2*diff_g)
            # releases
            for d in range(2, 2+num_bombs):
                r1 = random.random(); r2 = random.random()
                v[d] = (w*v[d] + c1*r1*(Pbest[i][d]-x[d]) + c2*r2*(Gbest[d]-x[d]))
            # fuses
            for d in range(2+num_bombs, 2+2*num_bombs):
                r1 = random.random(); r2 = random.random()
                v[d] = (w*v[d] + c1*r1*(Pbest[i][d]-x[d]) + c2*r2*(Gbest[d]-x[d]))
            # Clamp velocities
            v[0] = clamp(v[0], -vmax_speed, vmax_speed)
            v[1] = clamp(v[1], -vmax_heading, vmax_heading)
            for d in range(2, 2+num_bombs):
                v[d] = clamp(v[d], -vmax_tr, vmax_tr)
            for d in range(2+num_bombs, 2+2*num_bombs):
                v[d] = clamp(v[d], -vmax_tf, vmax_tf)
            # Update position and project
            x = [x[d] + v[d] for d in range(dims)]
            x = project_params(x)
            X[i] = x
            V[i] = v
            # Evaluate and update bests
            val = obj(x)
            if val > PbestVal[i]:
                Pbest[i] = x[:]
                PbestVal[i] = val
                if val > GbestVal:
                    GbestVal = val
                    Gbest = x[:]
    # Unpack best
    best_speed = Gbest[0]
    best_heading = Gbest[1]
    best_releases = Gbest[2:2+num_bombs]
    best_fuses = Gbest[2+num_bombs:2+2*num_bombs]
    return (best_speed, best_heading, best_releases, best_fuses),